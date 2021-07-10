from django.shortcuts import render
from django.http import HttpResponseRedirect
import psycopg2
import openpyxl
import xlrd
import os
import datetime

# Create your views here.
from Excel.forms import DocumentoForm
from Excel.models import Documento


def guardar_archivo(file):
    with open(file.name, "wb+") as destino:
        for chunk in file.chunks():
            destino.write(chunk)

def convertirFecha(num, book):
    return datetime.datetime(*xlrd.xldate_as_tuple(num, book.datemode))
def guardar_a_bd(file):
    
    try:
        conn = psycopg2.connect(database="postgres", user = "postgres", password = "postgres", host="db")
        cur = conn.cursor()
        print("Base abierta")
        
        if file.name.endswith(".xls"):
            table_name = file.name[:len(file.name) - 4].replace(" ", "_")
            
            cur.execute(f"SELECT EXISTS (SELECT * FROM pg_catalog.pg_tables WHERE schemaname != 'pg_catalog' AND schemaname != 'information_schema' AND tablename='{table_name.lower()}');")
            
            flag = bool(cur.fetchone()[0])

            if not flag:
                create = f"CREATE TABLE {table_name} ("
                insert = f"INSERT INTO {table_name} ("
                
                workbook = xlrd.open_workbook(file.name)
                ws = workbook.sheet_by_index(0)
                
                cabeceros = ws.row(0)
                aux = ws.row(1)
                
                for i in range(ws.ncols):
                    cadena = ws.cell_value(0,i).replace(" ", "_")
                    create += f"{cadena} "
                    
                    if(i == ws.ncols - 1):
                        if ws.cell_type(1,i) == 1:
                            create += "varchar"
                        elif ws.cell_type(1,i) == 3:
                            create += "date"
                        elif ws.cell_type(1,i) == 2:
                            create += "numeric"
                            
                        insert += f"{cadena}"
                    else:
                        if ws.cell_type(1,i) == 1:
                            create += "varchar, "
                        elif ws.cell_type(1,i) == 3:
                            create += "date, "
                        elif ws.cell_type(1,i) == 2:
                            create += "numeric, "
                            
                        insert += f"{cadena}, "
                        
                create += ");"
                insert += ") VALUES "
                        
                print(create)
                
                cur.execute(create)
                conn.commit()
                
                for i in range(1, ws.nrows):
                    for j in range(ws.ncols):
                        
                        if j == 0:
                            insert += "("
                        
                        if i == ws.nrows -1:
                            if j == ws.ncols - 1:
                                if ws.cell_type(i,j) == 3:
                                    insert += f"'{convertirFecha(ws.cell_value(i,j), workbook)}'); "
                                else:
                                    insert += f"'{ws.cell_value(i,j)}'); "
                            else:
                                if ws.cell_type(i,j) == 3:
                                    insert += f"'{convertirFecha(ws.cell_value(i,j), workbook)}', "
                                else:
                                    insert += f"'{ws.cell_value(i,j)}', "
                        else:
                            if j == ws.ncols - 1:
                                if ws.cell_type(i,j) == 3:
                                    insert += f"'{convertirFecha(ws.cell_value(i,j), workbook)}'), "
                                else:
                                    insert += f"'{ws.cell_value(i,j)}'), "
                            else:
                                if ws.cell_type(i,j) == 3:
                                    insert += f"'{convertirFecha(ws.cell_value(i,j), workbook)}', "
                                else:
                                    insert += f"'{ws.cell_value(i,j)}', "
                
                print(insert)
                
                cur.execute(insert)
                conn.commit()
                
                if os.path.exists(file.name):
                    os.remove(file.name)

                return "cargado"
            else:
                if os.path.exists(file.name):
                    os.remove(file.name)
                    
                return "existe"
        elif file.name.endswith(".xlsx"):
            table_name = file.name[:len(file.name) - 5].replace(" ", "_")
            
            cur.execute(f"SELECT EXISTS (SELECT * FROM pg_catalog.pg_tables WHERE schemaname != 'pg_catalog' AND schemaname != 'information_schema' AND tablename='{table_name.lower()}');")
            
            flag = bool(cur.fetchone()[0])
            
            if not flag:
            
                create = f"CREATE TABLE {table_name} ("
                insert = f"INSERT INTO {table_name} ("
                
                workbook = openpyxl.load_workbook(file.name, read_only=True)
                ws = workbook.active
                
                cabeceros = list(ws.rows)[0]
                aux = list(ws.rows)[1]
                
                for i in range(len(cabeceros)):
                    cadena = cabeceros[i].value.replace(" ", "_")
                    create += f"{cadena} "
                    
                    if(i == len(cabeceros)-1):
                        if aux[i].number_format == "General":
                            create += "varchar"
                        elif aux[i].number_format == "mm-dd-yy":
                            create += "date"
                        
                        insert += f"{cadena}"
                    else:
                        if aux[i].number_format == "General":
                            create += "varchar, "
                        elif aux[i].number_format == "mm-dd-yy":
                            create += "date, "
                            
                        insert += f"{cadena}, "
                        
                create += ");"
                insert += ") VALUES "
                        
                print(create)
                
                cur.execute(create)
                conn.commit()
                
                aux = list(ws.rows)
                
                for i in range(1, len(aux)):
                    for j in range(len(aux[i])):
                        
                        if j == 0:
                            insert += "("
                        
                        if i == ws.max_row -1:
                            if j == ws.max_column - 1:
                                insert += f"'{aux[i][j].value}'); "
                            else:
                                insert += f"'{aux[i][j].value}', "
                        else:
                            if j == ws.max_column - 1:
                                insert += f"'{aux[i][j].value}'), "
                            else:
                                insert += f"'{aux[i][j].value}', "
                
                print(insert)
                
                cur.execute(insert)
                conn.commit()
                
                if os.path.exists(file.name):
                    os.remove(file.name)
                
                return "cargado"
            else:
                if os.path.exists(file.name):
                    os.remove(file.name)
                    
                return "existe"
        else:
            if os.path.exists(file.name):
                os.remove(file.name)
                
            return "equivocado"
    except Exception as e:
        print(f"Error {e}")
        
        if os.path.exists(file.name):
                os.remove(file.name)
    finally:
        conn.close
        

def principal(request):
    if request.method == "POST":
        form = DocumentoForm(request.POST, request.FILES)

        if form.is_valid():

            guardar_archivo(request.FILES["file"])
            
            value = guardar_a_bd(request.FILES["file"])
            
            if value == "cargado":
                return render(request, "exito.html", {"mensaje": "¡Excel cargado con exito!"})
            elif value == "existe":
                return render(request, "exito.html", {"mensaje": "Ya subiste ese archivo antes"})
            elif value == "equivocado":
                return render(request, "exito.html", {"mensaje": "Verifica que sea un archivo de Excel"})
    else:
        form = DocumentoForm()

    return render(request, "principal.html", {"form": form})

def exito(request):
    return render(request, "exito.html")

def info(request):
    return render(request, "info.html")

